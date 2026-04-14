# -*- coding: utf-8 -*-
import io
import base64
from datetime import datetime, date
from odoo import _, fields, models, api
from odoo.exceptions import UserError


class ek_wizard_filter_regimen(models.TransientModel):
    _name = "ek.wizard.filter.regimen"
    _description = "Wizard: Filter Regimen"

    ek_regimen_table_id = fields.Many2one(
        "ek.ship.registration",
        string="Regimen",
        domain="[('assigned_regimen_70', '=', True)]"
    )
    ek_container_ids = fields.Many2many(
        "ek.boats.information",
        "ek_boats_information_id",
        "ek_wizard_filter_regimen",
        string="Container",
        domain="[('ship_name_id', '=', ek_regimen_table_id)]"
    )
    ek_id_bls = fields.Many2many(
        "id.bl.70",
        "ek_id_bl_70",
        "ek_wizard_filter_regimen",
        string="ID BL",
        domain="[('journey_crew_id', 'in', ek_container_ids)]"
    )
    date_from = fields.Date(string="Fecha Desde")
    date_to = fields.Date(string="Fecha Hasta")
    stage_ids = fields.Many2many(
        'ek.l10n.stages.mixin',
        'ek_wizard_filter_regimen_stage_rel',
        'wizard_id', 'stage_id',
        string="Estados"
    )

    @api.onchange('ek_regimen_table_id')
    def _onchange_ek_regimen_table_id(self):
        # Solo limpia las selecciones dependientes, NO auto-puebla
        self.ek_container_ids = False
        self.ek_id_bls = False

    @api.onchange('ek_container_ids')
    def _onchange_ek_container_id(self):
        # Solo limpia los BLs dependientes, NO auto-puebla
        self.ek_id_bls = False

    def _get_effective_containers(self):
        """Vacío = todos los contenedores del régimen seleccionado."""
        if self.ek_container_ids:
            return self.ek_container_ids
        if self.ek_regimen_table_id:
            return self.ek_regimen_table_id.ek_boats_information_ids
        return self.env['ek.boats.information'].browse()

    def generate_data_kardex(self):
        if not self.ek_regimen_table_id:
            raise UserError(_("Debe seleccionar un régimen."))

        containers = self._get_effective_containers()
        data = []

        # ── Régimen 70: entradas ─────────────────────────────────────────────
        for container in containers:
            items = container.ek_produc_packages_goods_ids
            if self.ek_id_bls:
                bl_ids = self.ek_id_bls.ids
                items = items.filtered(lambda x: x.id_bl.id in bl_ids)
            if self.date_from:
                date_from = self.date_from
                items = items.filtered(
                    lambda x: x.date_request and x.date_request.date() >= date_from
                )
            if self.date_to:
                date_to = self.date_to
                items = items.filtered(
                    lambda x: x.date_request and x.date_request.date() <= date_to
                )
            for item in items:
                data.append({
                    'date_request': item.date_request,
                    'movement_date': item.create_date,
                    'regimen': "70",
                    'ship_registration_id': item.ek_ship_registration_id.id,
                    'boats_information_id': item.ek_boats_information_id.id,
                    'related_document': item.id_bl.id,
                    'tariff_item': (item.tariff_item or "") + (item.id_hs_copmt_cd or "") + (item.id_hs_spmt_cd or ""),
                    'product_description': item.ek_requerid_burden_inter_nac_id.id,
                    'max_regime_date': item.max_regime_date,
                    'unds_in': abs(item.quantity or 0),
                    'unds_out': 0.0,
                    'unit_cost': item.fob,
                    'customs_document': item.number_dai,
                })

        # ── Régimen 60: salidas ──────────────────────────────────────────────
        container_ids = containers.ids
        if container_ids:
            domain = [
                ('ek_operation_request_id.container_id', 'in', container_ids),
                ('ek_operation_request_id.has_confirmed_type', '=', True),
                ('ek_operation_request_id.use_in_regimen_60', '=', True),
                ('ek_operation_request_id.canceled_stage', '!=', True),
            ]
            if self.ek_id_bls:
                domain.append(('ek_operation_request_id.id_bl_70', 'in', self.ek_id_bls.ids))
            if self.date_from:
                domain.append(('ek_operation_request_id.movement_date', '>=', self.date_from))
            if self.date_to:
                domain.append(('ek_operation_request_id.movement_date', '<=', self.date_to))
            if self.stage_ids:
                domain.append(('ek_operation_request_id.stage_id', 'in', self.stage_ids.ids))

            for item in self.env["table.regimen.60"].search(domain):
                req = item.ek_operation_request_id
                data.append({
                    'date_request': req.create_date,
                    'movement_date': req.movement_date,
                    'regimen': "60",
                    'ship_registration_id': req.ek_ship_registration_id.id,
                    'boats_information_id': req.journey_crew_id.id,
                    'related_document': req.id_bl_70.id,
                    'tariff_item': (item.tariff_item or "") + (item.id_hs_copmt_cd or "") + (item.id_hs_spmt_cd or ""),
                    'product_description': item.ek_requerid_burden_inter_nac_id.id,
                    'max_regime_date': req.max_regime_date,
                    'unds_in': 0.0,
                    'unds_out': abs(item.quantity or 0),
                    'unit_cost': item.fob,
                    'customs_document': req.number_dae,
                })

        return self._process_kardex_balance(data)

    def _process_kardex_balance(self, raw_data):
        """Agrupa por contenedor, ordena cronológicamente y calcula el
        saldo corrido por (subpartida, descripción) dentro de cada contenedor.

        El resultado queda ordenado: primero todos los movimientos del
        contenedor A, luego del B, etc. — dentro de cada uno, por fecha.
        """
        def _to_dt(val):
            if val is None:
                return datetime.min
            if isinstance(val, datetime):
                return val
            if isinstance(val, date):
                return datetime.combine(val, datetime.min.time())
            return datetime.min

        # Agrupar por contenedor
        by_container = {}
        for mv in raw_data:
            cid = mv.get('boats_information_id') or 0
            by_container.setdefault(cid, []).append(mv)

        # Orden estable de contenedores (por nombre)
        container_order = sorted(
            by_container.keys(),
            key=lambda cid: (
                self.env['ek.boats.information'].browse(cid).name or ''
                if cid else ''
            )
        )

        result = []
        for cid in container_order:
            movements = by_container[cid]
            # Ordenar cronológicamente por fecha de movimiento (tiebreak: date_request)
            movements.sort(key=lambda m: (
                _to_dt(m.get('movement_date')),
                _to_dt(m.get('date_request')),
                0 if m.get('regimen') == '70' else 1,  # entradas antes que salidas en misma fecha
            ))

            # Saldo por clave (subpartida + descripción)
            balances = {}
            for mv in movements:
                key = (mv.get('tariff_item') or '', mv.get('product_description') or 0)
                prev = balances.get(key, 0.0)
                new_bal = prev + (mv.get('unds_in') or 0) - (mv.get('unds_out') or 0)
                balances[key] = new_bal
                mv['balance'] = new_bal
                result.append(mv)

        return result

    def _create_charging_records(self):
        data = self.generate_data_kardex()
        if not data:
            raise UserError(_("No se encontraron datos con los filtros seleccionados."))
        self.env["ek.wizard.charging.data"].search([]).unlink()
        return self.env["ek.wizard.charging.data"].with_context(
            tracking_disable=True, validation_skip=True
        ).create(data)

    def create_wizard_charging(self):
        """Abre la tabla interactiva para revisión."""
        self._create_charging_records()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Kardex Régimen 70 / 60',
            'res_model': 'ek.wizard.charging.data',
            'view_mode': 'tree',
            'view_id': self.env.ref(
                "ek_l10n_shipping_operations_charging_regimes.ek_wizard_charging_data_form"
            ).id,
            'target': 'new',
            'context': {'create': False, 'edit': False, 'delete': False},
        }

    def create_wizard_charging_pdf(self):
        """Genera el PDF resumen del kárdex (1 fila por ítem, con agregados)."""
        containers = self._get_effective_containers()
        if not containers:
            raise UserError(_("Seleccione al menos un contenedor / régimen."))

        Item = self.env['ek.product.packagens.goods']
        domain = [('ek_boats_information_id', 'in', containers.ids)]
        if self.ek_id_bls:
            domain.append(('id_bl', 'in', self.ek_id_bls.ids))
        if self.date_from:
            domain.append(('date_request', '>=', self.date_from))
        if self.date_to:
            domain.append(('date_request', '<=', self.date_to))

        items = Item.search(domain, order='ek_operation_request_id desc, id_bl, sequence, id')
        if not items:
            raise UserError(_("No hay datos para los filtros seleccionados."))

        return self.env.ref(
            'ek_l10n_shipping_operations_charging_regimes.action_report_kardex_70_60'
        ).with_context(
            kardex_period_from=self.date_from,
            kardex_period_to=self.date_to,
        ).report_action(items)

    # ═════════════════════════════════════════════════════════════════
    # XLSX KÁRDEX – FORMATO 2026 (INVENTARIO REG 70 → 60)
    # ═════════════════════════════════════════════════════════════════

    def _vessel_color(self, ship):
        """Devuelve el código hex (sin #) configurado en el buque.

        `ship` puede ser un recordset ek.ship.registration o un string con
        el nombre (para retrocompatibilidad — busca por nombre y cachea).
        """
        if not ship:
            return 'FFFFFF'
        if isinstance(ship, str):
            ship = self.env['ek.ship.registration'].search(
                [('name', '=', ship)], limit=1
            )
            if not ship:
                return 'FFFFFF'
        color = (ship.kardex_color or '').strip().lstrip('#')
        if len(color) == 6:
            return color.upper()
        return 'FFFFFF'

    def _get_vessels_legend(self):
        """Devuelve los buques con color definido para la leyenda del Excel."""
        return self.env['ek.ship.registration'].search([
            ('kardex_color', '!=', False),
            ('kardex_color', '!=', ''),
            ('kardex_color', '!=', '#FFFFFF'),
        ])

    def _collect_kardex_2026_rows(self):
        """Devuelve una lista de diccionarios (1 por renglón del kárdex 2026).

        Por cada ítem importado (ek.product.packagens.goods) se generan:
          - 1 renglón por cada salida Reg.60 vinculada (con datos del Reg.60)
          - Si queda saldo > 0, 1 renglón extra "PENDIENTE" al final del ítem
          - Si no hay ninguna salida, 1 solo renglón con Reg.60 vacío
        """
        containers = self._get_effective_containers()
        if not containers:
            return []

        Reg60 = self.env['table.regimen.60']
        rows = []

        for container in containers.sorted('name'):
            items = container.ek_produc_packages_goods_ids
            if self.ek_id_bls:
                bl_ids = self.ek_id_bls.ids
                items = items.filtered(lambda x: x.id_bl.id in bl_ids)
            if self.date_from:
                items = items.filtered(
                    lambda x: x.date_request and x.date_request.date() >= self.date_from
                )
            if self.date_to:
                items = items.filtered(
                    lambda x: x.date_request and x.date_request.date() <= self.date_to
                )

            for item in items.sorted(lambda i: (i.sequence or 0, i.id)):
                req70 = item.ek_operation_request_id
                exits = Reg60.search([
                    ('ek_product_packagens_goods_id', '=', item.id),
                    ('ek_operation_request_id.canceled_stage', '!=', True),
                ], order='create_date')

                qty_in = abs(item.quantity or 0.0)
                delivered_acc = 0.0

                base_r70 = {
                    'carga': (req70.name or '') if req70 else '',
                    'referencia': (getattr(req70, 'type_name', '') or '') if req70 else '',
                    'bl': item.id_bl.name if item.id_bl else '',
                    'container': container.name or '',
                    'tipo': (container.type_container_id.name or '') if hasattr(container, 'type_container_id') and container.type_container_id else '',
                    'sol_previa': (req70.ek_assignment_number_rp or '') if req70 else '',
                    'dai': item.number_dai or '',
                    'buque_dest': item.ship_id.name if item.ship_id else '',
                    '_buque_dest_obj': item.ship_id,
                    'bultos': item.packages_count or '',
                    'item_prec': item.sequence or item.id,
                    'detalle': item.name or '',
                    'cantidad': qty_in,
                    'fecha_ingreso': req70.date_start if req70 else False,
                    'fecha_vto': item.max_regime_date,
                    'matricula': (req70.boat_registration or '') if req70 else '',
                    'observacion': item.observation or '',
                }

                if not exits:
                    # Ítem sin salidas → 1 renglón "EN STOCK"
                    rows.append({
                        **base_r70,
                        'buque_real': '', '_buque_real_obj': None,
                        'viaje': '', 'fc_almacopio': '', 'fc_nsk': '',
                        'dae': '', 'fecha_salida': False, 'cant_entregada': 0,
                        'fecha_entrega': False, 'destino': '', 'bl_exp': '',
                        'dae_reg': False, 'sustitutiva': '', 'dias_deposito': 0,
                        'saldo': qty_in,
                        'estado': self._compute_state(qty_in, item.max_regime_date),
                    })
                    continue

                for idx, ex in enumerate(exits):
                    req60 = ex.ek_operation_request_id
                    qty_out = abs(ex.quantity or 0.0)
                    delivered_acc += qty_out
                    saldo = qty_in - delivered_acc
                    ship_real = req60.ek_ship_registration_id if req60 else None
                    rows.append({
                        **base_r70,
                        'buque_real': (ship_real.name or '') if ship_real else '',
                        '_buque_real_obj': ship_real,
                        'viaje': (req60.booking_number or '') if req60 else '',
                        'fc_almacopio': (req60.delivery_report_code or '') if req60 else '',
                        'fc_nsk': (req60.delivery_report_sequence_code or '') if req60 else '',
                        'dae': (req60.number_dae or '') if req60 else '',
                        'fecha_salida': req60.atd if req60 else False,
                        'cant_entregada': qty_out,
                        'fecha_entrega': req60.date_end if req60 else False,
                        'destino': (ship_real.name or '') if ship_real else '',
                        'bl_exp': (req60.bl_import_export_id2.name or '') if req60 and req60.bl_import_export_id2 else '',
                        'dae_reg': bool(req60.dae_regularized) if req60 else False,
                        'sustitutiva': '',
                        'dias_deposito': (req60.days_in_deposit or 0) if req60 else 0,
                        'saldo': saldo,
                        'estado': self._compute_state(saldo, item.max_regime_date),
                    })

                # Si queda saldo > 0 después de todas las salidas → renglón PENDIENTE
                saldo_final = qty_in - delivered_acc
                if saldo_final > 0.0001:
                    rows.append({
                        **base_r70,
                        'buque_real': '', '_buque_real_obj': None,
                        'viaje': '', 'fc_almacopio': '', 'fc_nsk': '',
                        'dae': '', 'fecha_salida': False, 'cant_entregada': 0,
                        'fecha_entrega': False, 'destino': '', 'bl_exp': '',
                        'dae_reg': False, 'sustitutiva': '', 'dias_deposito': 0,
                        'saldo': saldo_final,
                        'estado': self._compute_state(saldo_final, item.max_regime_date),
                        '_pending': True,
                    })

        return rows

    def _compute_state(self, saldo, max_regime_date):
        """REEXPORTADO / PARCIAL / PRÓX. VENCER / VENCIDO / EN STOCK."""
        if saldo <= 0.0001:
            return 'REEXPORTADO'
        today = fields.Date.context_today(self)
        if max_regime_date:
            exp = max_regime_date.date() if hasattr(max_regime_date, 'date') else max_regime_date
            delta = (exp - today).days
            if delta < 0:
                return 'VENCIDO'
            if delta <= 15:
                return 'PRÓX. VENCER'
        return 'PARCIAL'

    def create_wizard_charging_xlsx(self):
        """Genera el XLSX del Kárdex Reg. 70/60 con formato 2026."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        rows = self._collect_kardex_2026_rows()
        if not rows:
            raise UserError(_("No se encontraron datos con los filtros seleccionados."))

        wb = Workbook()
        ws = wb.active
        ws.title = "INVENTARIO REG 70-60"

        # ── Estilos ──────────────────────────────────────────────────────────
        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(bold=True, color='FFFFFF', size=9)
        hdr_fill_r70 = PatternFill('solid', fgColor='1F4E79')  # azul oscuro
        hdr_fill_r60 = PatternFill('solid', fgColor='C65911')  # naranja oscuro
        hdr_fill_ctl = PatternFill('solid', fgColor='375623')  # verde oscuro
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')
        title_font = Font(bold=True, size=14, color='1F4E79')

        # Colores estado
        state_fills = {
            'REEXPORTADO':  PatternFill('solid', fgColor='C6EFCE'),  # verde
            'PARCIAL':      PatternFill('solid', fgColor='FFEB9C'),  # amarillo
            'PRÓX. VENCER': PatternFill('solid', fgColor='FFC7AA'),  # naranja
            'VENCIDO':      PatternFill('solid', fgColor='FFC7CE'),  # rojo
            'EN STOCK':     PatternFill('solid', fgColor='D9E1F2'),  # azul claro
        }

        # ── Definición de columnas (32 totales) ──────────────────────────────
        # (header, width, key, align, number_format)
        columns = [
            # --- RÉGIMEN 70 (A-O) ---
            ('CARGAS REG. 70',      14, 'carga',        center, None),
            ('REFERENCIA',          14, 'referencia',   left,   None),
            ('BL',                  18, 'bl',           left,   None),
            ('SERIE CONT./SUELTA',  18, 'container',    left,   None),
            ('TIPO',                 8, 'tipo',         center, None),
            ('SOLICITUD PREVIA',    14, 'sol_previa',   center, None),
            ('DAI',                 22, 'dai',          left,   None),
            ('BUQUE',               16, 'buque_dest',   center, None),
            ('BULTOS',               8, 'bultos',       center, None),
            ('ITEM PRECEDENTE',     10, 'item_prec',    center, None),
            ('DETALLE DE ITEMS',    40, 'detalle',      left,   None),
            ('CANTIDAD',            10, 'cantidad',     right,  '#,##0.00'),
            ('FECHA DE INGRESO',    13, 'fecha_ingreso',center, 'dd/mm/yyyy'),
            ('FECHA DE VENCIMIENTO',13, 'fecha_vto',    center, 'dd/mm/yyyy'),
            ('MATRICULA',           12, 'matricula',    center, None),
            # --- RÉGIMEN 60 (P-AB) ---
            ('BUQUE',               16, 'buque_real',   center, None),
            ('# VIAJE',             11, 'viaje',        center, None),
            ('FC. INFORMATIVA ALMACOPIO', 16, 'fc_almacopio', center, None),
            ('FC. EXP. NSK',        16, 'fc_nsk',       center, None),
            ('DAE',                 22, 'dae',          left,   None),
            ('FECHA SALIDA DEPÓSITO', 13, 'fecha_salida', center, 'dd/mm/yyyy'),
            ('CANTIDAD ENTREGADA',  12, 'cant_entregada', right, '#,##0.00'),
            ('FECHA ENTREGA AL BARCO', 13, 'fecha_entrega', center, 'dd/mm/yyyy'),
            ('DESTINO FINAL',       16, 'destino',      center, None),
            ('OBSERVACIONES',       25, 'observacion',  left,   None),
            ('BL EXP',              18, 'bl_exp',       left,   None),
            ('DAE REGULARIZADA',    12, 'dae_reg',      center, None),
            ('SUSTITUTIVA/CORRECTIVA/HOJA CAMBIO', 16, 'sustitutiva', center, None),
            # --- CONTROL / SALDO (AC-AF) ---
            ('DÍAS EN DEPÓSITO',    10, 'dias_deposito', center, '#,##0'),
            ('SALDO',               10, 'saldo',        right,  '#,##0.00'),
            ('ESTADO CONTENEDOR',   14, 'estado',       center, None),
            ('FACTURADO',           10, 'facturado',    center, None),
        ]

        ncols = len(columns)
        r70_end = 15   # col O
        r60_end = 28   # col AB

        # ── Filas 1-9: Cabecera institucional + leyenda ──────────────────────
        company = self.env.company
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=f"KARDEX INVENTARIO RÉGIMEN 70 → 60  ·  {company.name}")
        c.font = title_font
        c.alignment = center
        ws.row_dimensions[1].height = 22

        user_name = self.env.user.name
        period = ''
        if self.date_from or self.date_to:
            period = f"Período: {self.date_from or '...'} al {self.date_to or '...'}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=2, column=1, value=f"Generado: {fields.Datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Usuario: {user_name}  ·  {period}").alignment = center

        # Leyenda de colores por buque (leída desde ek.ship.registration.kardex_color)
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)
        ws.cell(row=4, column=1, value='LEYENDA DE COLORES POR BUQUE:').font = Font(bold=True)
        legend_row = 5
        col_pos = 1
        for vessel in self._get_vessels_legend():
            color = (vessel.kardex_color or '').lstrip('#').upper()
            if len(color) != 6:
                continue
            label = vessel.kardex_color_name and f"{vessel.name} ({vessel.kardex_color_name})" or vessel.name
            cell = ws.cell(row=legend_row, column=col_pos, value=label)
            cell.fill = PatternFill('solid', fgColor=color)
            cell.alignment = center
            cell.border = border
            col_pos += 1
            if col_pos > ncols:
                break

        # ── Fila 7: Encabezados agrupados (merge) ────────────────────────────
        group_row = 7
        ws.merge_cells(start_row=group_row, start_column=1, end_row=group_row, end_column=r70_end)
        g1 = ws.cell(row=group_row, column=1, value='RÉGIMEN 70 - IMPORTACIÓN')
        g1.font = hdr_font
        g1.fill = hdr_fill_r70
        g1.alignment = hdr_align
        g1.border = border

        ws.merge_cells(start_row=group_row, start_column=r70_end + 1, end_row=group_row, end_column=r60_end)
        g2 = ws.cell(row=group_row, column=r70_end + 1, value='RÉGIMEN 60 - REEXPORTACIÓN')
        g2.font = hdr_font
        g2.fill = hdr_fill_r60
        g2.alignment = hdr_align
        g2.border = border

        ws.merge_cells(start_row=group_row, start_column=r60_end + 1, end_row=group_row, end_column=ncols)
        g3 = ws.cell(row=group_row, column=r60_end + 1, value='CONTROL / SALDO')
        g3.font = hdr_font
        g3.fill = hdr_fill_ctl
        g3.alignment = hdr_align
        g3.border = border
        ws.row_dimensions[group_row].height = 22

        # ── Fila 8: Encabezados individuales ─────────────────────────────────
        hdr_row = 8
        for col_idx, (header, width, key, align, _nf) in enumerate(columns, 1):
            cell = ws.cell(row=hdr_row, column=col_idx, value=header)
            cell.font = hdr_font
            if col_idx <= r70_end:
                cell.fill = hdr_fill_r70
            elif col_idx <= r60_end:
                cell.fill = hdr_fill_r60
            else:
                cell.fill = hdr_fill_ctl
            cell.alignment = hdr_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[hdr_row].height = 32

        # Freeze panes y autofiltro
        ws.freeze_panes = f'A{hdr_row + 1}'
        ws.auto_filter.ref = f'A{hdr_row}:{get_column_letter(ncols)}{hdr_row}'

        # ── Filas de datos ───────────────────────────────────────────────────
        current_row = hdr_row + 1
        total_qty = 0.0
        total_delivered = 0.0
        total_saldo = 0.0

        for row_data in rows:
            for col_idx, (_hdr, _w, key, align, num_fmt) in enumerate(columns, 1):
                val = row_data.get(key, '')
                # Normalizar valores
                if key == 'facturado':
                    val = ''
                if key == 'dae_reg':
                    val = 'SÍ' if val else ''
                if isinstance(val, bool):
                    val = 'SÍ' if val else ''
                cell = ws.cell(row=current_row, column=col_idx, value=val if val != '' else None)
                cell.alignment = align
                cell.border = border
                if num_fmt:
                    cell.number_format = num_fmt

            # Color de fondo del buque destinatario (col H = 8)
            vessel_color = self._vessel_color(row_data.get('_buque_dest_obj'))
            if vessel_color != 'FFFFFF':
                ws.cell(row=current_row, column=8).fill = PatternFill('solid', fgColor=vessel_color)
            # Color de fondo del buque real (col P = 16)
            vessel_real_color = self._vessel_color(row_data.get('_buque_real_obj'))
            if vessel_real_color != 'FFFFFF':
                ws.cell(row=current_row, column=16).fill = PatternFill('solid', fgColor=vessel_real_color)

            # Color de estado (col AE = 31)
            estado = row_data.get('estado', '')
            if estado in state_fills:
                state_cell = ws.cell(row=current_row, column=31)
                state_cell.fill = state_fills[estado]
                state_cell.font = Font(bold=True, size=9)

            # Fila "pendiente" (sin salidas Reg60) en itálica
            if row_data.get('_pending'):
                for col_idx in range(16, r60_end + 1):
                    ws.cell(row=current_row, column=col_idx).font = Font(italic=True, color='808080')

            total_qty += row_data.get('cantidad', 0) or 0
            total_delivered += row_data.get('cant_entregada', 0) or 0
            current_row += 1

        # Saldo total = suma distinta (por ítem), no por renglón
        total_saldo = total_qty - total_delivered

        # ── Fila de totales ──────────────────────────────────────────────────
        totals_row = current_row + 1
        tfill = PatternFill('solid', fgColor='1F4E79')
        tfont = Font(bold=True, color='FFFFFF', size=10)
        ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=11)
        tcell = ws.cell(row=totals_row, column=1, value='TOTALES GENERALES')
        tcell.font = tfont
        tcell.fill = tfill
        tcell.alignment = right
        tcell.border = border

        ws.cell(row=totals_row, column=12, value=total_qty).number_format = '#,##0.00'
        ws.cell(row=totals_row, column=22, value=total_delivered).number_format = '#,##0.00'
        ws.cell(row=totals_row, column=30, value=total_saldo).number_format = '#,##0.00'
        for col in (12, 22, 30):
            c2 = ws.cell(row=totals_row, column=col)
            c2.font = tfont
            c2.fill = tfill
            c2.border = border
            c2.alignment = right
        ws.row_dimensions[totals_row].height = 22

        # ── Guardar y descargar ──────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_b64 = base64.b64encode(buf.getvalue()).decode()

        ship_name = (self.ek_regimen_table_id.name or 'TODOS').replace(' ', '_')
        fname_date = fields.Date.context_today(self).strftime('%Y%m%d')
        attachment = self.env['ir.attachment'].create({
            'name': f"INVENTARIO_REG_70_60_{ship_name}_{fname_date}.xlsx",
            'type': 'binary',
            'datas': xlsx_b64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
