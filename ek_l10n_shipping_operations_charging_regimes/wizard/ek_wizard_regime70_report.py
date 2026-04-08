# -*- coding: utf-8 -*-
import io
import base64
from odoo import _, api, fields, models
from odoo.exceptions import UserError


class EkWizardRegime70Report(models.TransientModel):
    _name = 'ek.wizard.regime70.report'
    _description = 'Wizard: Reportes Régimen 70 / 60 / Pólizas'

    report_type = fields.Selection([
        ('regime_70', 'Régimen 70 - Cargas en Depósito'),
        ('regime_60', 'Régimen 60 - Salidas'),
        ('insurance', 'Pólizas de Transporte'),
    ], string='Tipo de Reporte', required=True, default='regime_70')

    ship_id = fields.Many2one(
        'ek.ship.registration',
        string='Buque / Regimen',
        domain="[('assigned_regimen_70', '=', True)]"
    )
    container_ids = fields.Many2many(
        'ek.boats.information',
        'ek_wizard_r70_report_container_rel',
        'wizard_id', 'container_id',
        string='Contenedores',
        domain="[('ship_name_id', '=', ship_id)]"
    )
    date_from = fields.Date(string='Fecha Desde')
    date_to = fields.Date(string='Fecha Hasta')
    stage_ids = fields.Many2many(
        'ek.l10n.stages.mixin',
        'ek_wizard_r70_report_stage_rel',
        'wizard_id', 'stage_id',
        string='Estados'
    )

    @api.onchange('ship_id')
    def _onchange_ship_id(self):
        # Solo limpia — no auto-puebla. Vacío = todos.
        self.container_ids = False

    # ─── Domain builders ─────────────────────────────────────────────────────
    def _apply_common_filters(self, domain, date_field='date_transfer'):
        if self.container_ids:
            domain.append(('container_id', 'in', self.container_ids.ids))
        elif self.ship_id:
            domain.append(('ek_ship_registration_id', '=', self.ship_id.id))
        if self.date_from:
            domain.append((date_field, '>=', self.date_from))
        if self.date_to:
            domain.append((date_field, '<=', self.date_to))
        if self.stage_ids:
            domain.append(('stage_id', 'in', self.stage_ids.ids))
        return domain

    def _get_regime70_domain(self):
        domain = [('regime', '=', '70'), ('canceled_stage', '!=', True)]
        return self._apply_common_filters(domain, 'date_transfer')

    def _get_regime60_domain(self):
        domain = [
            ('use_in_regimen_60', '=', True),
            ('has_confirmed_type', '=', True),
            ('canceled_stage', '!=', True),
        ]
        return self._apply_common_filters(domain, 'movement_date')

    def _get_insurance_domain(self):
        domain = [('regime', '=', '70'), ('mail_sent_insurance', '=', True)]
        return self._apply_common_filters(domain, 'date_transfer')

    def _get_records(self):
        Req = self.env['ek.operation.request']
        if self.report_type == 'regime_70':
            return Req.search(self._get_regime70_domain()), 'Régimen 70'
        if self.report_type == 'regime_60':
            return Req.search(self._get_regime60_domain()), 'Régimen 60'
        return Req.search(self._get_insurance_domain()), 'Pólizas'

    # ─── PDF ─────────────────────────────────────────────────────────────────
    def action_generate_pdf(self):
        self.ensure_one()
        records, _label = self._get_records()
        if not records:
            raise UserError(_('No se encontraron solicitudes con los filtros seleccionados.'))
        ref_map = {
            'regime_70': 'ek_l10n_shipping_operations_charging_regimes.action_report_regime_70_deposit_cargo',
            'regime_60': 'ek_l10n_shipping_operations_charging_regimes.action_report_regime_60_exits',
            'insurance': 'ek_l10n_shipping_operations_charging_regimes.action_report_regime_70_insurance_applications',
        }
        return self.env.ref(ref_map[self.report_type]).report_action(records)

    # Backwards-compat alias
    def action_generate_report(self):
        return self.action_generate_pdf()

    # ─── XLSX ────────────────────────────────────────────────────────────────
    def action_generate_xlsx(self):
        self.ensure_one()
        records, label = self._get_records()
        if not records:
            raise UserError(_('No se encontraron solicitudes con los filtros seleccionados.'))

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = label[:31]

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(bold=True, color="FFFFFF", size=9)
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        grp_fill = PatternFill("solid", fgColor="2E75B6")
        title_font = Font(bold=True, size=13, color="1F4E79")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def fmt_date(val):
            if not val:
                return ''
            try:
                return val.strftime('%d/%m/%Y')
            except Exception:
                return str(val)[:10]

        def fmt_date_long(val):
            if not val:
                return ''
            try:
                dias = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']
                meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                         'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
                return f"{dias[val.weekday()]}, {val.day} de {meses[val.month-1]} de {val.year}"
            except Exception:
                return str(val)[:10]

        # ══════════════════════════════════════════════════════════════════
        if self.report_type == 'regime_70':
            ship_name = (self.ship_id and self.ship_id.name) or 'TODOS'
            year = (self.date_from and self.date_from.year) or fields.Date.today().year

            # Título
            ws.merge_cells('A1:R1')
            t = ws.cell(row=1, column=1, value=f"REGIMEN 70 - {ship_name} - {year}")
            t.font = title_font
            t.alignment = left
            ws.row_dimensions[1].height = 22

            ws.merge_cells('A2:R2')
            ws.cell(row=2, column=1,
                    value="Reporte de cargas ingresadas a depósito e importaciones.")

            # Cabecera agrupada (fila 4 grupo, fila 5 detalle)
            # Grupos: cols 6-9 IMPORTACIÓN, cols 12-16 EXPORTACIÓN
            ws.merge_cells('F4:I4')
            g1 = ws.cell(row=4, column=6, value='IMPORTACIÓN')
            g1.font = hdr_font
            g1.fill = grp_fill
            g1.alignment = hdr_align
            g1.border = border

            ws.merge_cells('L4:P4')
            g2 = ws.cell(row=4, column=12, value='EXPORTACIÓN')
            g2.font = hdr_font
            g2.fill = grp_fill
            g2.alignment = hdr_align
            g2.border = border

            headers = [
                ("SOLICITUD PREVIA", 14),    # 1
                ("NAVIEP", 10),               # 2
                ("#CONTENEDOR", 14),          # 3
                ("DETALLE", 14),              # 4
                ("TIPO CARGA", 12),           # 5
                ("BL-AWB INGRESO", 18),       # 6
                ("DAI-IMPORT", 20),           # 7
                ("# MATRICULA", 12),          # 8
                ("FECHA INGRESO DEPÓSITO", 22),  # 9
                ("BUQUE", 22),                # 10
                ("FAC. EXP DEPÓSITO", 16),    # 11
                ("DAE - EXP", 18),            # 12
                ("FECHA SALIDA DEPÓSITO", 22),   # 13
                ("BL EXPORTACIÓN", 20),       # 14
                ("DAE REGULARIZADA", 14),     # 15
                ("DÍAS EN DEPÓSITO", 12),     # 16
                ("ESTADO CONTENEDOR", 18),    # 17
                ("FACTURA ALMACENERA", 18),   # 18
            ]
            hdr_row = 5
            for col, (h, w) in enumerate(headers, 1):
                c = ws.cell(row=hdr_row, column=col, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align
                c.border = border
                ws.column_dimensions[c.column_letter].width = w
            ws.row_dimensions[hdr_row].height = 32

            def row_values(r):
                return [
                    r.name or '',
                    (r.shipping_lines and r.shipping_lines.name) or '',
                    (r.container_id and r.container_id.name) or '',
                    r.detail_supplies_spare_parts or '',
                    dict(r._fields['type_move_fcl_lcl'].selection).get(r.type_move_fcl_lcl, '') if r.type_move_fcl_lcl else '',
                    r.number_bl or '',
                    r.number_dai or '',
                    r.authorization_code or '',
                    fmt_date_long(r.date_transfer),
                    r.ek_ship_registration_id.name or '',
                    '',  # FAC. EXP DEPÓSITO (sin campo directo)
                    r.number_dae or '',
                    fmt_date_long(r.movement_date),
                    (r.bl_import_export_id and r.bl_import_export_id.name) or '',
                    'SI' if r.dae_regularized else 'NO',
                    r.days_in_deposit or 0,
                    r.stage_id.name or '',
                    '',  # FACTURA ALMACENERA (sin campo directo)
                ]

            data_start = hdr_row + 1
            for idx, rec in enumerate(records, data_start):
                for col, val in enumerate(row_values(rec), 1):
                    c = ws.cell(row=idx, column=col, value=val)
                    c.border = border
                    c.alignment = left if col in (2, 4, 10) else center

        # ══════════════════════════════════════════════════════════════════
        elif self.report_type == 'regime_60':
            ship_name = (self.ship_id and self.ship_id.name) or 'TODOS'
            year = (self.date_from and self.date_from.year) or fields.Date.today().year

            ws.merge_cells('A1:L1')
            t = ws.cell(row=1, column=1, value=f"REGIMEN 60 - {ship_name} - {year}")
            t.font = title_font
            t.alignment = left
            ws.row_dimensions[1].height = 22

            ws.merge_cells('A2:L2')
            ws.cell(row=2, column=1,
                    value="Reporte de salidas de depósito (exportaciones).")

            headers = [
                ("SOLICITUD", 14),
                ("NAVIEP", 10),
                ("#CONTENEDOR", 14),
                ("DETALLE", 16),
                ("BUQUE DESTINO", 22),
                ("DAE - EXP", 18),
                ("FECHA SALIDA DEPÓSITO", 22),
                ("BL EXPORTACIÓN", 20),
                ("DAE REGULARIZADA", 14),
                ("DÍAS EN DEPÓSITO", 12),
                ("FECHA MAX RÉGIMEN", 16),
                ("ESTADO", 18),
            ]
            hdr_row = 4
            for col, (h, w) in enumerate(headers, 1):
                c = ws.cell(row=hdr_row, column=col, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align
                c.border = border
                ws.column_dimensions[c.column_letter].width = w
            ws.row_dimensions[hdr_row].height = 32

            def row_values(r):
                return [
                    r.name or '',
                    (r.shipping_lines and r.shipping_lines.name) or '',
                    (r.container_id and r.container_id.name) or '',
                    r.detail_supplies_spare_parts or '',
                    r.ek_ship_registration_id.name or '',
                    r.number_dae or '',
                    fmt_date_long(r.movement_date),
                    (r.bl_import_export_id and r.bl_import_export_id.name) or '',
                    'SI' if r.dae_regularized else 'NO',
                    r.days_in_deposit or 0,
                    fmt_date(r.max_regime_date),
                    r.stage_id.name or '',
                ]

            data_start = hdr_row + 1
            for idx, rec in enumerate(records, data_start):
                for col, val in enumerate(row_values(rec), 1):
                    c = ws.cell(row=idx, column=col, value=val)
                    c.border = border
                    c.alignment = left if col in (2, 4, 5) else center

        # ══════════════════════════════════════════════════════════════════
        else:  # insurance
            # Tomar datos de póliza del primer registro (típicamente todos comparten)
            first = records[0]
            poliza_madre = first.authorization_code or ''
            fecha_contrato = first.insurance_contract_date
            renov_desde = first.insurance_renewal_from
            renov_hasta = first.insurance_renewal_to

            # ── Cabecera tipo Zurich ────────────────────────────────────
            ws.merge_cells('A1:H1')
            t = ws.cell(row=1, column=1, value='DETALLE DE APLICACIONES - PÓLIZA DE TRANSPORTE')
            t.font = title_font
            t.alignment = left
            ws.row_dimensions[1].height = 22

            info_font = Font(bold=True, size=10)
            ws.cell(row=3, column=1, value='PÓLIZA MADRE:').font = info_font
            ws.cell(row=3, column=2, value=poliza_madre)

            ws.cell(row=4, column=1, value='FECHA DE CONTRATO:').font = info_font
            ws.cell(row=4, column=2, value=fmt_date_long(fecha_contrato))

            ws.cell(row=5, column=1, value='DESDE:').font = info_font
            ws.cell(row=5, column=2, value=fmt_date(fecha_contrato))
            ws.cell(row=5, column=3, value='HASTA:').font = info_font
            ws.cell(row=5, column=4, value=fmt_date(renov_desde))

            ws.cell(row=6, column=1, value='RENOVACIÓN DESDE:').font = info_font
            ws.cell(row=6, column=2, value=fmt_date(renov_desde))
            ws.cell(row=6, column=3, value='HASTA:').font = info_font
            ws.cell(row=6, column=4, value=fmt_date(renov_hasta))

            headers = [
                ("FECHA", 22),
                ("# APLICACIÓN", 12),
                ("AÑO", 8),
                ("ASEGURADO", 12),
                ("VÍA DE TRANSPORTE", 14),
                ("TIPO MERCADERÍA", 22),
                ("PTO. EMBARQUE", 14),
                ("PTO. DESTINO", 14),
                ("VALOR ASEGURADO", 14),
                ("VALOR DE FLETE", 14),
                ("FECHA DE EMBARQUE", 14),
                ("CONTENEDOR", 14),
                ("BUQUE", 20),
                ("BL", 18),
                ("# CERTIFICADO", 14),
                ("PRIMA", 12),
                ("FACTURA", 18),
            ]
            hdr_row = 8
            for col, (h, w) in enumerate(headers, 1):
                c = ws.cell(row=hdr_row, column=col, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align
                c.border = border
                ws.column_dimensions[c.column_letter].width = w
            ws.row_dimensions[hdr_row].height = 32

            def row_values(r):
                return [
                    fmt_date_long(r.mail_sent_insurance_date or r.date_transfer),
                    r.name or '',  # # aplicación
                    (r.date_transfer and r.date_transfer.year) or '',
                    'NSK',
                    'TERRESTRE',
                    r.detail_supplies_spare_parts or '',
                    '',  # PTO. EMBARQUE (sin campo directo)
                    '',  # PTO. DESTINO (sin campo directo)
                    '',  # VALOR ASEGURADO (sin campo directo)
                    '',  # VALOR DE FLETE (sin campo directo)
                    fmt_date(r.on_board_date),
                    (r.container_id and r.container_id.name) or '',
                    r.ek_ship_registration_id.name or '',
                    r.number_bl or '',
                    r.authorization_code or '',
                    '',  # PRIMA (sin campo directo)
                    '',  # FACTURA (sin campo directo)
                ]

            data_start = hdr_row + 1
            for idx, rec in enumerate(records, data_start):
                for col, val in enumerate(row_values(rec), 1):
                    c = ws.cell(row=idx, column=col, value=val)
                    c.border = border
                    c.alignment = left if col in (1, 4, 5, 6, 13) else center

        # Guardar y descargar
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_b64 = base64.b64encode(buf.getvalue()).decode()

        filename = f"{label.replace(' ', '_')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': xlsx_b64,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
