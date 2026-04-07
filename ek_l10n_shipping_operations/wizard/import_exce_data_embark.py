import base64
import io

import openpyxl
from odoo import _, fields, models
from odoo.exceptions import ValidationError


class ImportExcelDataEmbarkWizard(models.TransientModel):
  """
  Wizard para importar datos de embarque desde archivos Excel.

  Permite importar información de mercancías y paquetes desde archivos Excel
  hacia operaciones navales, con soporte para diferentes regímenes aduaneros.
  """

  _name = 'import.excel.data.embark.wizard'
  _description = 'Import Excel File'

  file = fields.Binary(
    string='Excel File',
    help='Archivo Excel con datos de embarque a importar (.xlsx)',
  )
  filename_document = fields.Char('File Name')

  ek_operation_request_id = fields.Many2one('ek.operation.request')
  type = fields.Selection(
    [('anexo1', 'Anexo1')], string='Type Document', default='anexo1'
  )

  use_in_regimen_61 = fields.Boolean(
    related='ek_operation_request_id.type_id.use_in_regimen_61'
  )

  def get_workbook(self, file_content):
    """
    Decodifica y carga un archivo Excel desde contenido base64.

    Args:
        file_content: Contenido del archivo Excel en base64

    Returns:
        openpyxl.Workbook: Libro de trabajo cargado

    Raises:
        ValidationError: Si hay errores al procesar el archivo
    """
    try:
      inputx = io.BytesIO()
      inputx.write(base64.decodebytes(file_content))
      inputx.seek(0)
      return openpyxl.load_workbook(inputx, data_only=True)
    except Exception as e:
      raise ValidationError('Error: {}'.format(e))

  def import_records_attendance(self):
    """
    Importa registros desde el archivo Excel.

    Maneja dos flujos principales:
    - Anexo1 sin régimen 61: Procesa y crea nuevos registros
    - Con régimen 61: Copia datos desde operación relacionada
    """
    self.ensure_one()
    if self.type == 'anexo1' and not self.use_in_regimen_61:
      self.ek_operation_request_id.ek_produc_packages_goods_ids.unlink()
      ids = []
      try:
        book = self.get_workbook(self.file)
        sheet = book.worksheets[0]
        values = self.process_xlsx_sheet(sheet)
        items_read = [x.get('name') for x in values[0]]
        items_exit = self.env['ek.requerid.burden.inter.nac'].search(
          [('name', 'in', items_read)]
        )
        dict_items = {x.name: x.id for x in items_exit}
        for vals in values[0]:
          data = vals.get('name')
          inter_nac_id = dict_items.get(data, False)
          if not inter_nac_id:
            _values = self.env['ek.requerid.burden.inter.nac'].create(
              {'name': data}
            )
            inter_nac_id = _values.id

          vals.update(
            {
              'ek_requerid_burden_inter_nac_id': inter_nac_id,
            }
          )

          # NUEVO: Buscar/crear producto automáticamente
          # El método create() del modelo ya se encarga de esto vía _find_or_create_product
          # pero mantenemos compatibilidad
          itens = self.env['ek.product.packagens.goods'].create(vals)
          ids.append(itens.id)
        self.file = False
      except Exception as e:
        raise ValidationError(_('Error importing XLSX file\n{}'.format(e)))
    if (
      self.use_in_regimen_61
      and self.ek_operation_request_id.ek_operation_request_id
    ):
      lines = []
      for line in self.ek_operation_request_id.ek_operation_request_id.ek_produc_packages_goods_ids:
        lines.append(
          (
            0,
            0,
            {
              'tariff_item': line.tariff_item,
              'id_hs_copmt_cd': line.id_hs_copmt_cd,
              'id_hs_spmt_cd': line.id_hs_spmt_cd,
              'ek_requerid_burden_inter_nac_id': line.ek_requerid_burden_inter_nac_id.id,
              'name': line.name,
              'quantity': line.quantity,
              'gross_weight': line.gross_weight,
              'product_weight_in_lbs': line.product_weight_in_lbs,
              'fob': line.fob,
              'invoice_number': line.invoice_number,
              'supplier': line.supplier,
            },
          )
        )
      self.ek_operation_request_id.ek_produc_packages_goods_ids = lines

  def get_spanish_headers(self):
    """
    Retorna los encabezados en español para que el usuario los use como plantilla.
    Separa claramente entre columnas obligatorias y opcionales.

    Returns:
        dict: Diccionario con columnas obligatorias y opcionales
    """
    return {
      'obligatorias': [
        'Descripción de Mercancías',      # OBLIGATORIO
        'Nombre del Proveedor',           # OBLIGATORIO  
        'Número de Factura',              # OBLIGATORIO
        'Cantidad de Empaques',           # OBLIGATORIO
      ],
      'opcionales': [
        'Código de Partida Arancelaria',  # OPCIONAL
        'Código HS Complementario',       # OPCIONAL
        'Código HS Suplementario',        # OPCIONAL
        'Peso Neto',                      # OPCIONAL
        'Valor FOB Unitario',             # OPCIONAL
      ]
    }

  def clean_header_text(self, text):
    """
    Limpia el texto del encabezado de caracteres especiales y espacios.

    Args:
        text: Texto a limpiar

    Returns:
        str: Texto limpio
    """
    if not text:
      return ''
    # Convertir a string si no lo es
    text = str(text)
    # Reemplazar caracteres especiales comunes
    text = text.replace('\xa0', ' ')  # Reemplazar non-breaking space
    text = text.replace('\u200b', '')  # Reemplazar zero-width space
    # Eliminar espacios al inicio y final
    text = text.strip()
    return text

  def process_xlsx_sheet(self, sheet):
    """
    Procesa una hoja de Excel y extrae los datos validados.

    Args:
        sheet: Hoja de Excel a procesar

    Returns:
        tuple: (valores_procesados, errores)

    Raises:
        ValidationError: Si hay errores de validación en los datos
    """
    values = []
    errors = ''

    # Mapeo de encabezados en español a códigos originales
    header_spanish_mapping = {
      'Código de Partida Arancelaria': 'icl_hs_part_cd',
      'Código HS Complementario': 'icl_hs_cpmt_cd',
      'Código HS Suplementario': 'icl_hs_spmt_cd',
      'Descripción de Mercancías': 'icl_gds_desc_cn',
      'Peso Neto': 'icl_net_wt',
      'Cantidad de Empaques': 'icl_phsc_pck_ut_co',
      'Valor FOB Unitario': 'icl_fobv_ut_pr',
      'Número de Factura': 'icl_item_inv_no',
      'Nombre del Proveedor': 'icl_pvdr_nm_01',
    }

    # Definir columnas obligatorias y opcionales
    required_columns = {
      'Descripción de Mercancías': 'icl_gds_desc_cn',
      'Nombre del Proveedor': 'icl_pvdr_nm_01',
      'Número de Factura': 'icl_item_inv_no',
      'Cantidad de Empaques': 'icl_phsc_pck_ut_co',
    }

    optional_columns = {
      'Código de Partida Arancelaria': 'icl_hs_part_cd',
      'Código HS Complementario': 'icl_hs_cpmt_cd',
      'Código HS Suplementario': 'icl_hs_spmt_cd',
      'Peso Neto': 'icl_net_wt',
      'Valor FOB Unitario': 'icl_fobv_ut_pr',
    }

    # Leer encabezado y limpiar caracteres especiales
    header = [
      self.clean_header_text(cell.value)
      for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]

    if not header or all(not h for h in header):
      raise ValidationError(
        _('El archivo Excel no tiene encabezados en la primera fila')
      )

    # Crear mapeo de índices
    header_mapping = {}
    found_columns = []
    missing_required_columns = []

    # Verificar columnas OBLIGATORIAS
    for spanish_name, internal_code in required_columns.items():
      found = False
      for idx, col in enumerate(header):
        clean_spanish_name = self.clean_header_text(spanish_name)
        clean_col = self.clean_header_text(col)

        if clean_col == clean_spanish_name:
          header_mapping[internal_code] = idx
          found_columns.append(spanish_name)
          found = True
          break
      if not found:
        missing_required_columns.append(spanish_name)

    # Si faltan columnas obligatorias, mostrar error
    if missing_required_columns:
      raise ValidationError(
        _(
          'Columnas OBLIGATORIAS faltantes en el archivo Excel: {}.\n\n'
          'Las columnas obligatorias son:\n{}\n\n'
          'Las columnas opcionales (pueden estar vacías) son:\n{}'
        ).format(
          ', '.join(missing_required_columns),
          '\n'.join('- ' + h for h in required_columns.keys()),
          '\n'.join('- ' + h for h in optional_columns.keys()),
        )
      )

    # Verificar columnas OPCIONALES (no generar error si faltan)
    for spanish_name, internal_code in optional_columns.items():
      found = False
      for idx, col in enumerate(header):
        clean_spanish_name = self.clean_header_text(spanish_name)
        clean_col = self.clean_header_text(col)

        if clean_col == clean_spanish_name:
          header_mapping[internal_code] = idx
          found_columns.append(spanish_name)
          found = True
          break
      # No agregar a missing_columns si es opcional

    # Validar y convertir valores numéricos
    def validate_numeric_value(
      value, field_name, row_number, max_value=999999999.99
    ):
      """
      Valida que el valor sea numérico y válido.

      Args:
          value: Valor a validar
          field_name: Nombre del campo para mensajes de error
          row_number: Número de fila para mensajes de error
          max_value: Valor máximo permitido

      Returns:
          float: Valor numérico validado

      Raises:
          ValidationError: Si el valor no es válido
      """
      if value is None or value == '':
        return 0.0

      try:
        # Convertir a float
        numeric_value = float(value)
        # Validar que no sea negativo
        if numeric_value < 0:
          raise ValidationError(
            _(
              'El valor de {} en la fila {} no puede ser negativo: {}'.format(
                field_name, row_number, value
              )
            )
          )
        # Validar que no sea excesivamente grande
        if numeric_value > max_value:
          raise ValidationError(
            _(
              'El valor de {} en la fila {} es demasiado grande. Máximo permitido: {}'.format(
                field_name, row_number, max_value
              )
            )
          )
        return numeric_value
      except (ValueError, TypeError):
        raise ValidationError(
          _(
            'El valor de {} en la fila {} debe ser numérico: {}'.format(
              field_name, row_number, value
            )
          )
        )

    # Validar campos requeridos
    def validate_required_field(value, field_name, row_number):
      """
      Valida que el campo requerido no esté vacío.

      Args:
          value: Valor a validar
          field_name: Nombre del campo para mensajes de error
          row_number: Número de fila para mensajes de error

      Returns:
          Valor validado

      Raises:
          ValidationError: Si el campo está vacío
      """
      if not value:
        raise ValidationError(
          _(
            'El campo {} en la fila {} es requerido y no puede estar vacío'.format(
              field_name, row_number
            )
          )
        )
      return value

    # Función para obtener valor de columna (opcional o requerida)
    def get_column_value(row, mapping_key, default_value=''):
      """Obtiene el valor de una columna si existe, sino retorna valor por defecto"""
      if mapping_key in header_mapping:
        return row[header_mapping[mapping_key]]
      return default_value

    # Procesar filas
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
      try:
        row = [cell.value for cell in row]
        if not any(row):  # Skip empty rows
          continue

        vals = {}

        # Campos OBLIGATORIOS - validar que existan y no estén vacíos
        vals['name'] = validate_required_field(
          get_column_value(row, 'icl_gds_desc_cn'),
          'Descripción de Mercancías',
          row_idx,
        )
        vals['supplier'] = validate_required_field(
          get_column_value(row, 'icl_pvdr_nm_01'),
          'Nombre del Proveedor',
          row_idx,
        )
        vals['invoice_number'] = validate_required_field(
          get_column_value(row, 'icl_item_inv_no'),
          'Número de Factura',
          row_idx,
        )
        vals['quantity'] = validate_numeric_value(
          validate_required_field(
            get_column_value(row, 'icl_phsc_pck_ut_co'),
            'Cantidad de Empaques',
            row_idx,
          ),
          'Cantidad de Empaques',
          row_idx,
        )

        # Campos OPCIONALES - usar valores por defecto si no existen
        vals['tariff_item'] = get_column_value(row, 'icl_hs_part_cd', '')
        vals['id_hs_copmt_cd'] = get_column_value(row, 'icl_hs_cpmt_cd', '')
        vals['id_hs_spmt_cd'] = get_column_value(row, 'icl_hs_spmt_cd', '')

        # Campos numéricos opcionales
        gross_weight_value = get_column_value(row, 'icl_net_wt', 0)
        vals['gross_weight'] = (
          validate_numeric_value(gross_weight_value, 'Peso Neto', row_idx)
          if gross_weight_value not in [None, '', 0]
          else 0.0
        )

        fob_value = get_column_value(row, 'icl_fobv_ut_pr', 0)
        vals['fob'] = (
          validate_numeric_value(
            fob_value,
            'Valor FOB Unitario',
            row_idx,
            max_value=999999999.99,
          )
          if fob_value not in [None, '', 0]
          else 0.0
        )

        vals['ek_operation_request_id'] = self.ek_operation_request_id.id
        values.append(vals)

      except ValidationError as ve:
        raise ValidationError(_('Error en la fila {}: {}'.format(row_idx, ve)))

    return values, errors
